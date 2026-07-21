import csv
from io import BytesIO, StringIO
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select

from app.database.session import get_db
from app.models.transaction import Transaction
from app.models.category import Category
from app.models.account import Account
from app.models.user import User
from app.routers.deps import get_current_user

router = APIRouter()

def _build_transaction_query(
    db: Session,
    user_id: int,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
    category_id: Optional[int],
    account_id: Optional[int],
    tx_type: Optional[str],
    transaction_ids: Optional[List[int]] = None
):
    """Build a filtered or id-specific transaction query for exports."""
    stmt = (
        select(Transaction)
        .options(
            joinedload(Transaction.category),
            joinedload(Transaction.account),
            joinedload(Transaction.to_account),
        )
        .where(Transaction.user_id == user_id)
        .order_by(Transaction.date.desc())
    )
    if transaction_ids:
        stmt = stmt.where(Transaction.id.in_(transaction_ids))
    else:
        if start_date:
            stmt = stmt.where(Transaction.date >= start_date)
        if end_date:
            stmt = stmt.where(Transaction.date <= end_date)
        if category_id:
            stmt = stmt.where(Transaction.category_id == category_id)
        if account_id:
            stmt = stmt.where(Transaction.account_id == account_id)
        if tx_type:
            stmt = stmt.where(Transaction.type == tx_type)
    return db.execute(stmt).scalars().all()


def _tx_to_row(tx):
    """Convert a transaction to a flat dict for export."""
    return {
        "Date": tx.date.strftime("%Y-%m-%d") if tx.date else "",
        "Type": (tx.type or "").capitalize(),
        "Description": tx.description or "",
        "Amount": f"{tx.amount:.2f}",
        "Category": tx.category.name if tx.category else "",
        "Source Account": tx.account.name if tx.account else "",
        "Destination Account": tx.to_account.name if tx.to_account else "",
    }


@router.get("/csv")
def export_csv(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    category_id: Optional[int] = None,
    account_id: Optional[int] = None,
    type: Optional[str] = None,
    ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export transactions as CSV download with optional filters or selection."""
    transactions = _build_transaction_query(
        db, current_user.id, start_date, end_date, category_id, account_id, type, ids
    )

    output = StringIO()
    fieldnames = ["Date", "Type", "Description", "Amount", "Category", "Source Account", "Destination Account"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for tx in transactions:
        writer.writerow(_tx_to_row(tx))

    output.seek(0)
    filename = f"expenseflow_export_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel")
def export_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    category_id: Optional[int] = None,
    account_id: Optional[int] = None,
    type: Optional[str] = None,
    ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export transactions as Excel (.xlsx) download with optional filters or selection."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl",
        )

    transactions = _build_transaction_query(
        db, current_user.id, start_date, end_date, category_id, account_id, type, ids
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = ["Date", "Type", "Description", "Amount", "Category", "Source Account", "Destination Account"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, tx in enumerate(transactions, 2):
        row_data = _tx_to_row(tx)
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = thin_border
            if key == "Amount":
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"

    col_widths = [12, 10, 30, 14, 18, 20, 20]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"expenseflow_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/pdf")
def export_pdf(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    category_id: Optional[int] = None,
    account_id: Optional[int] = None,
    type: Optional[str] = None,
    ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export transactions as PDF download with optional filters or selection."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF export requires reportlab. Install with: pip install reportlab",
        )

    transactions = _build_transaction_query(
        db, current_user.id, start_date, end_date, category_id, account_id, type, ids
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        textColor=colors.HexColor('#1E1B4B'),
        fontSize=20,
        spaceAfter=12
    )
    normal_style = ParagraphStyle(
        'DocNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )

    # Title & Metadata
    story.append(Paragraph("ExpenseFlow AI — Transaction Export", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} for {current_user.full_name or current_user.email}", normal_style))
    story.append(Spacer(1, 15))

    # Table Header & Data Rows
    headers = ["Date", "Type", "Description", "Amount", "Category", "Account"]
    data = [headers]

    for tx in transactions:
        row_data = _tx_to_row(tx)
        data.append([
            row_data["Date"],
            row_data["Type"],
            row_data["Description"][:25] + "..." if len(row_data["Description"]) > 25 else row_data["Description"],
            row_data["Amount"],
            row_data["Category"],
            row_data["Source Account"]
        ])

    col_widths = [65, 55, 150, 65, 100, 105]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # align amount right
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
    ])
    table.setStyle(t_style)
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    filename = f"expenseflow_export_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
